import sys
import sqlite3
import json
import os
from openpyxl import Workbook

from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QToolBar,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
    QHeaderView,
    QStatusBar,
    QFileDialog,
    QMessageBox,
    QAction,
    QLabel,
    QDialog,
    QLineEdit,
    QCheckBox,
    QDialogButtonBox,
    QFormLayout,
    QProgressBar
)
from PySide6.QtCore import QThread
from PySide6.QtGui import QIcon

from agent import OSMAgent


SETTINGS_FILE = "settings.json"


class SettingsDialog(QDialog):
    """Dialog for configuring application settings."""
    
    def __init__(self, parent=None, current_settings=None):
        super().__init__(parent)
        self.setWindowTitle("Settings")
        self.setModal(True)
        
        layout = QFormLayout(self)
        
        # Margo API Key field
        self.margo_key_input = QLineEdit()
        self.margo_key_input.setEchoMode(QLineEdit.Password)
        self.margo_key_input.setPlaceholderText("Optional – free key from margo.io")
        layout.addRow("Margo API Key:", self.margo_key_input)
        
        # Use Margo checkbox
        self.use_margo_checkbox = QCheckBox("Use Margo to enrich leads (20/day free)")
        layout.addRow(self.use_margo_checkbox)
        
        # Groq API Key field
        self.groq_key_input = QLineEdit()
        self.groq_key_input.setEchoMode(QLineEdit.Password)
        self.groq_key_input.setPlaceholderText("Optional – Groq API key for lead scoring")
        layout.addRow("Groq API Key:", self.groq_key_input)
        
        # Use Groq checkbox
        self.use_groq_checkbox = QCheckBox("Use Groq for lead scoring")
        layout.addRow(self.use_groq_checkbox)
        
        # Dialog buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addRow(button_box)
        
        # Load current settings
        if current_settings:
            self.margo_key_input.setText(current_settings.get('margo_key', ''))
            self.use_margo_checkbox.setChecked(current_settings.get('use_margo', False))
            self.groq_key_input.setText(current_settings.get('groq_key', ''))
            self.use_groq_checkbox.setChecked(current_settings.get('use_groq', False))
    
    def get_settings(self):
        """Return the settings as a dict."""
        return {
            'margo_key': self.margo_key_input.text(),
            'use_margo': self.use_margo_checkbox.isChecked(),
            'groq_key': self.groq_key_input.text(),
            'use_groq': self.use_groq_checkbox.isChecked()
        }


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        
        self.setWindowTitle("LeadFalcon – Leather Leads")
        self.resize(1000, 600)
        
        # Load settings from file
        self.app_settings = self._load_settings()
        
        # Create toolbar
        toolbar = QToolBar("Main Toolbar")
        self.addToolBar(toolbar)
        
        # Toolbar actions
        self.start_action = QAction("Start", self)
        self.start_action.triggered.connect(self.on_start)
        toolbar.addAction(self.start_action)
        
        self.pause_action = QAction("Pause", self)
        self.pause_action.triggered.connect(self.on_pause)
        toolbar.addAction(self.pause_action)
        
        self.stop_action = QAction("Stop", self)
        self.stop_action.triggered.connect(self.on_stop)
        toolbar.addAction(self.stop_action)
        
        self.export_action = QAction("Export", self)
        self.export_action.triggered.connect(self.on_export)
        toolbar.addAction(self.export_action)
        
        # Settings action with gear icon
        self.settings_action = QAction("Settings", self)
        try:
            # Try to use a standard gear icon if available
            self.settings_action.setIcon(QIcon.fromTheme("preferences-system"))
        except Exception:
            pass
        self.settings_action.triggered.connect(self.on_settings)
        toolbar.addAction(self.settings_action)
        
        # Central widget with table
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        layout = QVBoxLayout(central_widget)
        
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(7)
        self.table_widget.setHorizontalHeaderLabels([
            "Type", "Name", "Phone", "Email", "Website", "City", "Source"
        ])
        
        header = self.table_widget.horizontalHeader()
        header.setStretchLastSection(True)
        header.setSectionResizeMode(QHeaderView.Stretch)
        
        layout.addWidget(self.table_widget)
        
        # Progress bar (indeterminate during active scraping)
        self.progress_bar = QProgressBar()
        self.progress_bar.setMaximum(0)  # Indeterminate mode
        self.progress_bar.hide()
        layout.addWidget(self.progress_bar)
        
        # Progress label for lead counter
        self.progress_label = QLabel("Leads: 0")
        layout.addWidget(self.progress_label)
        
        # Status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("Ready")
        
        # Keep references to thread and agent
        self.thread = None
        self.agent = None
        
        # Lead counter
        self.lead_count = 0
    
    def _load_settings(self):
        """Load settings from settings.json file."""
        if os.path.exists(SETTINGS_FILE):
            try:
                with open(SETTINGS_FILE, 'r') as f:
                    return json.load(f)
            except Exception as e:
                print(f"Error loading settings: {e}")
        return {}
    
    def _save_settings(self, settings):
        """Save settings to settings.json file."""
        try:
            with open(SETTINGS_FILE, 'w') as f:
                json.dump(settings, f, indent=2)
            return True
        except Exception as e:
            print(f"Error saving settings: {e}")
            return False
    
    def on_settings(self):
        """Open the settings dialog."""
        dialog = SettingsDialog(self, self.app_settings)
        if dialog.exec():
            new_settings = dialog.get_settings()
            self.app_settings = new_settings
            self._save_settings(new_settings)
            self.status_bar.showMessage("Settings saved")

    def on_lead_found(self, lead: dict):
        """Add a new lead to the table."""
        row_position = self.table_widget.rowCount()
        self.table_widget.insertRow(row_position)
        
        self.table_widget.setItem(row_position, 0, QTableWidgetItem(lead.get("type", "")))
        self.table_widget.setItem(row_position, 1, QTableWidgetItem(lead.get("name", "")))
        self.table_widget.setItem(row_position, 2, QTableWidgetItem(lead.get("phone", "")))
        self.table_widget.setItem(row_position, 3, QTableWidgetItem(lead.get("email", "")))
        self.table_widget.setItem(row_position, 4, QTableWidgetItem(lead.get("website", "")))
        self.table_widget.setItem(row_position, 5, QTableWidgetItem(lead.get("city", "")))
        self.table_widget.setItem(row_position, 6, QTableWidgetItem(lead.get("source", "")))
        
        self.table_widget.scrollToBottom()
        
        # Update lead counter
        self.lead_count += 1
        self.progress_label.setText(f"Leads: {self.lead_count}")
        self.status_bar.showMessage(f"Total leads: {self.lead_count}")

    def on_start(self):
        """Start the agent in a new thread."""
        if self.thread is not None and self.thread.isRunning():
            self.status_bar.showMessage("Agent already running")
            return
        
        self.thread = QThread()
        self.agent = OSMAgent(settings=self.app_settings)
        
        self.agent.moveToThread(self.thread)
        
        # Connect signals
        self.thread.started.connect(self.agent.run)
        self.agent.finished.connect(self.thread.quit)
        self.agent.finished.connect(self.agent.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)
        self.agent.finished.connect(self.on_agent_finished)
        
        self.agent.status_updated.connect(self.status_bar.showMessage)
        self.agent.lead_found.connect(self.on_lead_found)
        
        # Show progress bar during scraping
        self.progress_bar.show()
        
        self.thread.start()
        self.status_bar.showMessage("Agent starting...")
    
    def on_agent_finished(self):
        """Called when agent finishes processing."""
        self.progress_bar.hide()
        self.status_bar.showMessage(f"All cities processed. Total leads: {self.lead_count}")

    def on_pause(self):
        """Pause the agent."""
        if self.agent:
            self.agent.pause()

    def on_stop(self):
        """Stop the agent and thread."""
        if self.agent:
            self.agent.stop()
        if self.thread:
            self.thread.quit()
            self.thread.wait()
            self.thread = None
            self.agent = None
        self.status_bar.showMessage("Agent stopped")

    def on_export(self):
        """Export table data to Excel."""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to Excel",
            "",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Leads"
            
            # Write headers
            headers = ["Type", "Name", "Phone", "Email", "Website", "City", "Source"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write data
            for row in range(self.table_widget.rowCount()):
                for col in range(self.table_widget.columnCount()):
                    item = self.table_widget.item(row, col)
                    value = item.text() if item else ""
                    ws.cell(row=row + 2, column=col + 1, value=value)
            
            wb.save(file_path)
            self.status_bar.showMessage(f"Exported to {file_path}")
            QMessageBox.information(self, "Export Complete", f"Data exported to {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export: {e}")


if __name__ == "__main__":
    import database
    app = QApplication(sys.argv)
    database.initialize_db()
    
    # Check if cities table is empty and comuni.csv exists
    cities = database.get_cities()
    if not cities and os.path.exists(database.CSV_PATH):
        # File exists but no cities loaded (empty or malformed)
        pass  # Let the app start, user will see no data
    elif not cities and not os.path.exists(database.CSV_PATH):
        # File doesn't exist - show message to user
        from PySide6.QtWidgets import QMessageBox
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Warning)
        msg_box.setWindowTitle("Missing Data")
        msg_box.setText("comuni.csv not found – please run generate_cities.py first")
        msg_box.exec()
    
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
